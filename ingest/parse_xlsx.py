"""
ingest/parse_xlsx.py
--------------------
Parses PGDBA Batch XLSX files into the same chunk structure as parse.py (PDF parser).

XLSX format (Batch 5 and similar):
  Sheet "Student-wise":
    Row 1 — group headers: PERSONAL DETAILS | COMPANY-1 | COMPANY-2 | COMPANY-3
    Row 2 — column names
    Row 3+ — one student per row, up to 3 company experiences across columns

Each student × company pair → one chunk (identical structure to PDF chunks).
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import COMPANIES, COMPANY_ALIASES, extract_batch_from_filename
from ingest.parse import detect_topics, build_embed_text

# ── Column layout ─────────────────────────────────────────────────────────────
# Each block maps field → column index. Repeats 3x for Company-1/2/3.
COMPANY_BLOCKS = [
    {"company":10,"role":11,"process":12,"areas":13,
     "tech_rounds":14,"tech_q":15,"hr_q":16,"final":17},
    {"company":18,"role":19,"process":21,"areas":22,
     "tech_rounds":23,"tech_q":24,"hr_q":25,"final":26},
    {"company":27,"role":28,"process":30,"areas":31,
     "tech_rounds":32,"tech_q":33,"hr_q":34,"final":35},
]


def _s(val) -> str:
    return "" if val is None else str(val).strip()


def _normalise_company(raw: str) -> str:
    """Resolve aliases then match against known companies list."""
    upper = raw.upper().strip()
    # direct alias match
    if upper in COMPANY_ALIASES:
        return COMPANY_ALIASES[upper]
    # partial match against known list
    for known in COMPANIES:
        if known in upper or upper in known:
            return known
    return upper


def _build_full_text(row, block) -> str:
    parts = []
    for label, col in [
        ("Role",               block["role"]),
        ("Selection process",  block["process"]),
        ("Areas covered",      block["areas"]),
        ("Technical questions",block["tech_q"]),
        ("HR questions",       block["hr_q"]),
        ("Final round",        block["final"]),
    ]:
        val = _s(row[col])
        if val:
            parts.append(f"{label}:\n{val}")
    return "\n\n".join(parts)


def parse_xlsx(xlsx_path: Path, batch: int) -> list[dict]:
    wb = load_workbook(str(xlsx_path), read_only=True)

    sheet_name = "Student-wise"
    if sheet_name not in wb.sheetnames:
        # Fallback: try first non-empty sheet
        sheet_name = wb.sheetnames[0]
        print(f"  'Student-wise' not found, using '{sheet_name}'")

    ws = wb[sheet_name]
    all_rows = [r for r in ws.iter_rows(values_only=True)
                if any(v is not None for v in r)]
    data_rows = all_rows[2:]  # skip 2 header rows

    chunks = []
    for row_idx, row in enumerate(data_rows):
        # pad to safe length
        row = tuple(row) + (None,) * max(0, 36 - len(row))

        for b_idx, block in enumerate(COMPANY_BLOCKS):
            company_raw = _s(row[block["company"]])
            if not company_raw:
                continue

            company   = _normalise_company(company_raw)
            full_text = _build_full_text(row, block)
            if len(full_text) < 50:
                continue

            topics = detect_topics(full_text)

            # Build rounds from structured columns
            rounds = []
            for rname, col in [("Technical Round", block["tech_q"]),
                                ("HR Round",        block["hr_q"]),
                                ("Final Round",     block["final"])]:
                content = _s(row[col])
                if content:
                    rounds.append({"round_name": rname,
                                   "content":    content,
                                   "topics":     detect_topics(content)})
            if not rounds:
                rounds = [{"round_name":"general","content":full_text,"topics":topics}]

            chunk = {
                "chunk_id":   f"{company.lower().replace(' ','_')}_xlsx_b{batch}_{row_idx}_{b_idx}",
                "company":    company,
                "batch":      batch,
                "source":     xlsx_path.name,
                "topics":     topics,
                "rounds":     rounds,
                "num_rounds": len(rounds),
                "embed_text": build_embed_text(company, batch, topics, rounds, full_text),
                "full_text":  full_text,
            }
            chunks.append(chunk)

    return chunks
